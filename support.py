# import openpyxl module 
import openpyxl
import os
from datetime import date


today = date.today()
# dd/mm/YY
d1 = today.strftime("%d_%m")

#checking if result exist then delete it
file_path = d1+".xlsx"
if os.path.isfile(file_path):
  os.remove(file_path)
  print("File has been deleted")
else:
  print("File does not exist")


# Give the location of the file 
path = "clear_result.xlsx"
kod_path =  "Kodiak.xlsx"  

# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
wb_kod = openpyxl.load_workbook(kod_path)

# Get workbook active sheet object 
# from the active attribute 
clear_ws_obj = wb_obj.active
kod_ws_obj = wb_kod["3. Public Corporate Subscriber"]


clear_row = clear_ws_obj.max_row
clear_column = clear_ws_obj.max_column

print("Total Rows:", clear_row)
print("Total Columns:", clear_column)

#preparing result sheet header
result_sheet = wb_obj.create_sheet("result")
result_ws_obj = wb_obj["result"]
result_ws_obj.cell(row=1,column=1).value = "testcase"
result_ws_obj.cell(row=1,column=2).value = "testcase procedure"
result_ws_obj.cell(row=1,column=3).value = "remark"
result_ws_obj.cell(row=1,column=4).value = "screen shot"
result_ws_obj.cell(row=1,column=5).value = "trx"
result_ws_obj.cell(row=1,column=6).value = "code line"

# wb_obj.save(d1+".xlsx")
# print("result saved")

#adding data in result
result_row_index=2
for i in range(1, clear_row + 1): 
    cell_obj = clear_ws_obj.cell(row = i, column = 2)
    if (cell_obj.value == "Failed"):
        # print(cell_obj.value, end = " ")
        result_ws_obj.cell(row=result_row_index,column=1).value = clear_ws_obj.cell(row = i,column=1).value
        result_row_index = result_row_index + 1 
print("result row index : "+str(result_row_index))

kod_row = kod_ws_obj.max_row
print("kodrow"+str(kod_row))
def search_kodiak(testcase):
    print("in function\n")
    for i in range(3, kod_row + 1): 
        cell_obj1 = kod_ws_obj.cell(row = i, column = 1) 
        # print(cell_obj1.value+"\n")
        if(cell_obj1.value == testcase):
            print("success\n")
            return i
    return 0

new_result_row_index=2
#writing procedure
for i in range(2, result_row_index): 
    cell_obj = result_ws_obj.cell(row = i, column = 1)
    index = search_kodiak(cell_obj.value)
    # print(str(index) +" "+str(i)+cell_obj.value)
    if(index != 0):
        result_ws_obj.cell(row=new_result_row_index,column=2).value = kod_ws_obj.cell(row = index, column = 4).value
        new_result_row_index = new_result_row_index + 1


wb_obj.save(d1+".xlsx")
print("result saved")
